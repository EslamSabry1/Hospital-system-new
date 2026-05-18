# devices/views.py
import logging
from datetime import date, timedelta

from django.contrib import messages
from django.contrib.auth import get_user_model, login, logout
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db.models import Avg, Count, ExpressionWrapper, F, IntegerField, Q, Sum
from django.db.models.functions import Now
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET

import io
import qrcode
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .forms import DepartmentForm, DeviceForm, LoginForm, MaintenanceForm
from .models import Department, Device, Maintenance, MaintenanceTask, TechnicianNote
from .scheduling import sync_calendar
from .utils.prediction import compute_failure_prediction

logger = logging.getLogger(__name__)


# ─── JSON / API views ─────────────────────────────────────────────────────────

@require_GET
@login_required
def control_center_stats_api(request):
    today = timezone.now().date()
    total_devices = Device.objects.count()
    active_devices = Device.objects.filter(status="active").count()
    maintenance_devices = Device.objects.filter(status="maintenance").count()

    critical_alerts = Device.objects.filter(next_maintenance__isnull=False).filter(
        Q(next_maintenance__lt=today) | Q(next_maintenance__lte=today + timedelta(days=3))
    ).count()

    recent_qs = Maintenance.objects.select_related("device").order_by("-date")[:5]
    recent_list = [
        {
            "device_name": m.device.name if m.device else "",
            "device_id": m.device.device_id if m.device else "",
            "technician": m.technician or "Technician",
            "maintenance_type": m.get_maintenance_type_display(),
            "cost": float(m.cost or 0),
            "date": m.date.strftime("%b %d, %Y") if m.date else "",
        }
        for m in recent_qs
    ]

    system_health = round((active_devices / total_devices) * 100, 1) if total_devices else 0

    return JsonResponse({
        "counts": {
            "total_devices": total_devices,
            "active_devices": active_devices,
            "maintenance_devices": maintenance_devices,
            "critical_alerts": critical_alerts,
        },
        "system_health": system_health,
        "recent_maintenance": recent_list,
        "server_time": timezone.now().strftime("%H:%M:%S"),
    })


@require_GET
@login_required
def device_lookup_api(request):
    device_id = (request.GET.get("device_id") or "").strip()
    if not device_id:
        return JsonResponse({"ok": False, "error": "device_id is required"}, status=400)

    device = Device.objects.filter(device_id__iexact=device_id).first()
    if not device:
        return JsonResponse({"ok": False, "error": "Device not found"}, status=404)

    return JsonResponse({
        "ok": True,
        "pk": device.pk,
        "url": reverse("device_detail", kwargs={"pk": device.pk}),
    })


@require_GET
@login_required
def maintenance_calendar_api(request):
    sync_calendar(horizon_days=180)
    today = timezone.now().date()

    tasks = MaintenanceTask.objects.select_related("device", "template").order_by("due_date")[:200]
    events = [
        {
            "id": task.id,
            "title": f"{task.device.name} • {task.template.name}",
            "date": task.due_date.isoformat(),
            "status": task.status,
            "urgency": task.urgency,
            "reminder_date": task.reminder_date.isoformat(),
            "is_overdue": task.due_date < today and task.status != "completed",
        }
        for task in tasks
    ]

    return JsonResponse({"events": events, "count": len(events)})


# ─── Excel export ─────────────────────────────────────────────────────────────

@login_required
def devices_export_excel(request):
    qs = Device.objects.all()

    search = request.GET.get("search")
    status = request.GET.get("status")
    department = request.GET.get("department")

    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(device_id__icontains=search))
    if status:
        qs = qs.filter(status=status)
    if department:
        qs = qs.filter(department_id=department)

    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"

    headers = ["Device ID", "Name", "Type", "Department", "Status", "Model", "Serial"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    for d in qs:
        ws.append([
            d.device_id,
            d.name,
            d.get_device_type_display(),
            d.department.name if d.department else "",
            d.get_status_display(),
            d.model,
            d.serial_number,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=devices.xlsx"
    wb.save(response)
    return response


# ─── Auth ─────────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect("control_center")

    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get("username")
            user_model = get_user_model()
            user, _ = user_model.objects.get_or_create(username=username)
            login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            messages.success(request, "Logged in successfully!")
            return redirect("control_center")
        else:
            messages.error(request, "Invalid username or password")
    else:
        form = LoginForm()

    return render(request, "auth/login.html", {"form": form})


def logout_view(request):
    logout(request)
    messages.success(request, "Logged out successfully!")
    return redirect("login")


# ─── Dashboard ────────────────────────────────────────────────────────────────

def build_dashboard_context():
    today = timezone.now().date()
    total_devices = Device.objects.count()
    active_devices = Device.objects.filter(status="active").count()
    maintenance_devices = Device.objects.filter(status="maintenance").count()
    inactive_devices = Device.objects.filter(status="inactive").count()

    def percentage(count):
        return (count / total_devices * 100) if total_devices else 0

    upcoming_cutoff = today + timedelta(days=30)
    upcoming_maintenance = (
        Device.objects.select_related("department")
        .filter(next_maintenance__isnull=False, next_maintenance__lte=upcoming_cutoff)
        .order_by("next_maintenance", "name")
    )

    return {
        "total_devices": total_devices,
        "active_devices": active_devices,
        "maintenance_devices": maintenance_devices,
        "inactive_devices": inactive_devices,
        "total_departments": Department.objects.count(),
        "active_percentage": percentage(active_devices),
        "maintenance_percentage": percentage(maintenance_devices),
        "inactive_percentage": percentage(inactive_devices),
        "recent_devices": Device.objects.select_related("department").order_by("-created_at")[:5],
        "upcoming_maintenance": upcoming_maintenance,
        "today": today,
    }


@login_required
def dashboard(request):
    return render(request, "dashboard.html", build_dashboard_context())


# ─── Device CRUD ──────────────────────────────────────────────────────────────

@login_required
def device_list(request):
    devices_base = Device.objects.select_related("department").order_by("-created_at")

    status_filter = request.GET.get("status")
    department_filter = request.GET.get("department")
    search_query = request.GET.get("search")

    if department_filter:
        devices_base = devices_base.filter(department_id=department_filter)
    if search_query:
        devices_base = devices_base.filter(
            Q(name__icontains=search_query)
            | Q(device_id__icontains=search_query)
            | Q(serial_number__icontains=search_query)
        )

    devices = devices_base
    if status_filter:
        devices = devices.filter(status=status_filter)

    departments = Department.objects.all()

    base_params = request.GET.copy()
    if "status" in base_params:
        del base_params["status"]

    clear_status_url = f"?{base_params.urlencode()}" if base_params.urlencode() else "?"

    status_kpis = []
    for value, label in Device.DEVICE_STATUS:
        params = base_params.copy()
        params["status"] = value
        status_kpis.append({
            "value": value,
            "label": label,
            "count": devices_base.filter(status=value).count(),
            "url": f"?{params.urlencode()}",
            "is_selected": status_filter == value,
        })

    return render(request, "devices/list.html", {
        "devices": devices,
        "departments": departments,
        "status_choices": Device.DEVICE_STATUS,
        "status_kpis": status_kpis,
        "total_devices": devices_base.count(),
        "clear_status_url": clear_status_url,
    })


@login_required
def device_add(request):
    if request.method == "POST":
        form = DeviceForm(request.POST, request.FILES)
        if form.is_valid():
            device = form.save()
            try:
                device.generate_qr_code()
                device.save()
                messages.info(request, "QR Code generated successfully!")
            except Exception as exc:
                messages.warning(request, f"Device saved but QR code generation failed: {exc}")
            messages.success(request, f"Device {device.name} added successfully!")
            return redirect("device_detail", pk=device.pk)
    else:
        form = DeviceForm()

    return render(request, "devices/add_edit.html", {"form": form, "title": "Add New Device"})


@login_required
def device_edit(request, pk):
    device = get_object_or_404(Device, pk=pk)

    if request.method == "POST":
        form = DeviceForm(request.POST, request.FILES, instance=device)
        if form.is_valid():
            form.save()
            if "generate_qr" in request.POST:
                try:
                    device.generate_qr_code()
                    device.save()
                    messages.success(request, "QR Code regenerated successfully!")
                except Exception as exc:
                    messages.error(request, f"Failed to generate QR code: {exc}")
            messages.success(request, f"Device {device.name} updated successfully!")
            return redirect("device_list")
    else:
        form = DeviceForm(instance=device)

    return render(request, "devices/add_edit.html", {"form": form, "device": device, "title": "Edit Device"})


@login_required
def device_delete(request, pk):
    device = get_object_or_404(Device, pk=pk)
    if request.method == "POST":
        name = device.name
        device.delete()
        messages.success(request, f"Device {name} deleted successfully!")
        return redirect("device_list")
    return render(request, "devices/delete_confirm.html", {"device": device})


@login_required
def generate_device_qr(request, pk):
    device = get_object_or_404(Device, pk=pk)
    try:
        device.generate_qr_code()
        device.save()
        messages.success(request, "QR Code generated successfully!")
    except Exception as exc:
        messages.error(request, f"Failed to generate QR code: {exc}")
    return redirect("device_detail", pk=pk)


def device_detail(request, pk):
    device = get_object_or_404(Device, pk=pk)
    maintenances = device.maintenances.all().order_by("-date")
    is_qr_access = request.GET.get("qr") in ["1", "true", "True"]
    prediction = compute_failure_prediction(device, maintenances)

    if not request.user.is_authenticated:
        if request.method == "POST":
            messages.warning(request, "Please login to add maintenance records")
            return redirect("login")
        context = {
            "device": device,
            "maintenances": maintenances,
            "is_public": True,
            "is_qr_access": is_qr_access,
            "prediction": prediction,
        }
        if is_qr_access:
            return render(request, "devices/device_public_qr.html", context)
        return render(request, "devices/device_public.html", context)

    if request.method == "POST":
        maintenance_form = MaintenanceForm(request.POST, request.FILES)
        if maintenance_form.is_valid():
            maintenance = maintenance_form.save(commit=False)
            maintenance.device = device
            maintenance.save()
            device.last_maintenance = maintenance.date
            if maintenance.next_maintenance_date:
                device.next_maintenance = maintenance.next_maintenance_date
            device.save()
            messages.success(request, "Maintenance record added successfully!")
            return redirect("device_detail", pk=pk)
    else:
        maintenance_form = MaintenanceForm()

    return render(request, "devices/detail.html", {
        "device": device,
        "maintenances": maintenances,
        "maintenance_form": maintenance_form,
        "is_qr_access": is_qr_access,
        "prediction": prediction,
    })


def device_qr(request, pk):
    device = get_object_or_404(Device, pk=pk)
    device_path = reverse("device_detail", kwargs={"pk": device.pk})
    full_url = request.build_absolute_uri(f"{device_path}?qr=1")
    img = qrcode.make(full_url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return HttpResponse(buf.getvalue(), content_type="image/png")


# ─── Department CRUD ──────────────────────────────────────────────────────────

@login_required
def departments_list(request):
    departments = Department.objects.annotate(device_count=Count("device")).order_by("name")
    total_devices = sum(dept.device_count for dept in departments)
    avg_devices_per_dept = (total_devices / len(departments)) if departments else 0
    top_department = max(departments, key=lambda d: d.device_count) if departments else None

    for dept in departments:
        dept_devices = Device.objects.filter(department=dept)
        dept.active_devices = dept_devices.filter(status="active").count()
        dept.maintenance_devices = dept_devices.filter(status="maintenance").count()
        dept.inactive_devices = dept_devices.filter(status="inactive").count()
        if not hasattr(dept, "manager"):
            dept.manager = None

    return render(request, "devices/departments.html", {
        "departments": departments,
        "total_devices": total_devices,
        "avg_devices_per_dept": round(avg_devices_per_dept, 1),
        "top_department": top_department,
        "department_count": departments.count(),
    })


@login_required
def department_add(request):
    if request.method == "POST":
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save()
            messages.success(request, f"Department {department.name} added successfully!")
            return redirect("departments_list")
    else:
        form = DepartmentForm()
    return render(request, "devices/department_form.html", {"form": form, "title": "Add New Department"})


@login_required
def department_edit(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == "POST":
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, f"Department {department.name} updated successfully!")
            return redirect("departments_list")
    else:
        form = DepartmentForm(instance=department)
    return render(request, "devices/department_form.html", {
        "form": form, "department": department, "title": "Edit Department"
    })


@login_required
def department_delete(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == "POST":
        name = department.name
        department.delete()
        messages.success(request, f"Department {name} deleted successfully!")
        return redirect("departments_list")
    return render(request, "devices/department_delete.html", {"department": department})


# ─── Control Center ───────────────────────────────────────────────────────────

@login_required
def control_center(request):
    device_qs = Device.objects.select_related("department")

    selected_department = request.GET.get("department", "").strip()
    selected_device_type = request.GET.get("device_type", "").strip()
    selected_status = request.GET.get("status", "").strip()

    if selected_department:
        device_qs = device_qs.filter(department_id=selected_department)
    if selected_device_type:
        device_qs = device_qs.filter(device_type=selected_device_type)
    if selected_status:
        device_qs = device_qs.filter(status=selected_status)

    total_devices = device_qs.count()
    active_devices = device_qs.filter(status="active").count()
    maintenance_devices = device_qs.filter(status="maintenance").count()
    inactive_devices = device_qs.filter(status="inactive").count()

    today = date.today()
    critical_alerts = 0
    overdue_maintenance = []

    for device in device_qs.filter(next_maintenance__isnull=False):
        if device.next_maintenance:
            delta = device.next_maintenance - today
            if delta.days < 0:
                critical_alerts += 1
                overdue_maintenance.append({
                    "name": device.name,
                    "device_id": device.device_id,
                    "days_overdue": abs(delta.days),
                    "next_maintenance": device.next_maintenance,
                })
            elif delta.days <= 3:
                critical_alerts += 1

    system_health = min((active_devices / total_devices) * 100, 100) if total_devices else 0

    return render(request, "control_center.html", {
        "total_devices": total_devices,
        "active_devices": active_devices,
        "maintenance_devices": maintenance_devices,
        "inactive_devices": inactive_devices,
        "total_departments": Department.objects.count(),
        "critical_alerts": critical_alerts,
        "devices_normal": active_devices,
        "devices_pending": maintenance_devices,
        "devices_critical": critical_alerts,
        "overdue_maintenance": overdue_maintenance,
        "has_critical_alerts": critical_alerts > 0,
        "system_health": system_health,
        "mean_downtime": "2.3 hours",
        "avg_repair_time": "4.7 hours",
        "recent_maintenance": Maintenance.objects.select_related("device").order_by("-date")[:5],
        "todays_maintenance": Maintenance.objects.select_related("device").filter(date=today).order_by("-created_at")[:5],
        "devices_due_today": device_qs.filter(next_maintenance=today).count(),
        "departments": Department.objects.order_by("name"),
        "device_types": Device.DEVICE_TYPE,
        "statuses": Device.DEVICE_STATUS,
        "selected_department": selected_department,
        "selected_device_type": selected_device_type,
        "selected_status": selected_status,
    })


# ─── Reports ──────────────────────────────────────────────────────────────────

@login_required
def reports_view(request):
    today = date.today()
    total_devices = Device.objects.count()
    active_devices = Device.objects.filter(status="active").count()
    maintenance_devices = Device.objects.filter(status="maintenance").count()
    inactive_devices = Device.objects.filter(status="inactive").count()

    overdue_maintenance = list(
        Device.objects.filter(next_maintenance__lt=today, next_maintenance__isnull=False)
    )
    for d in overdue_maintenance:
        d.days_overdue = (today - d.next_maintenance).days

    maintenance_stats = Maintenance.objects.aggregate(
        total_cost=Sum("cost"), avg_cost=Avg("cost"), total_records=Count("id")
    )

    thirty_days_ago = today - timedelta(days=30)
    recent_maintenance = Maintenance.objects.filter(date__gte=thirty_days_ago).order_by("-date")

    departments = Department.objects.annotate(
        device_count=Count("device"),
        active_count=Count("device", filter=Q(device__status="active")),
        maintenance_count=Count("device", filter=Q(device__status="maintenance")),
        inactive_count=Count("device", filter=Q(device__status="inactive")),
    ).order_by("-device_count")

    device_types = Device.objects.values("device_type").annotate(
        count=Count("id"),
        active=Count("id", filter=Q(status="active")),
        maintenance=Count("id", filter=Q(status="maintenance")),
        inactive=Count("id", filter=Q(status="inactive")),
    ).order_by("-count")

    maintenance_types = Maintenance.objects.values("maintenance_type").annotate(
        count=Count("id"), avg_cost=Avg("cost"), total_cost=Sum("cost")
    ).order_by("-count")

    monthly_trend = []
    for i in range(5, -1, -1):
        month_start = today.replace(day=1) - timedelta(days=30 * i)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        agg = Maintenance.objects.filter(date__range=[month_start, month_end]).aggregate(
            count=Count("id"), total_cost=Sum("cost"), avg_cost=Avg("cost")
        )
        monthly_trend.append({
            "month": month_start.strftime("%b %Y"),
            "count": agg["count"] or 0,
            "cost": agg["total_cost"] or 0,
            "avg_cost": agg["avg_cost"] or 0,
        })

    top_technicians = Maintenance.objects.values("technician").annotate(
        count=Count("id"), avg_cost=Avg("cost"), total_cost=Sum("cost")
    ).order_by("-count")[:5]

    warranty_active = Device.objects.filter(warranty_expiry__gte=today).count()
    warranty_expired = Device.objects.filter(warranty_expiry__lt=today).count()

    return render(request, "reports.html", {
        "total_devices": total_devices,
        "active_devices": active_devices,
        "maintenance_devices": maintenance_devices,
        "inactive_devices": inactive_devices,
        "critical_devices": len(overdue_maintenance),
        "maintenance_total_cost": maintenance_stats["total_cost"] or 0,
        "maintenance_avg_cost": maintenance_stats["avg_cost"] or 0,
        "maintenance_total_records": maintenance_stats["total_records"] or 0,
        "overdue_maintenance": overdue_maintenance,
        "has_overdue_maintenance": bool(overdue_maintenance),
        "departments": departments,
        "total_departments": departments.count(),
        "device_types": device_types,
        "maintenance_types": maintenance_types,
        "monthly_trend": monthly_trend,
        "top_technicians": top_technicians,
        "warranty_active": warranty_active,
        "warranty_expired": warranty_expired,
        "warranty_percentage": min(
            max((warranty_active / total_devices * 100) if total_devices else 0, 0), 100
        ),
        "recent_maintenance": recent_maintenance[:10],
        "today": today,
        "thirty_days_ago": thirty_days_ago,
        "currency_symbol": "EGP",
        "currency_code": "EGP",
    })


# ─── Procurement ──────────────────────────────────────────────────────────────

@login_required
def procurement_dashboard(request):
    devices = Device.objects.prefetch_related("maintenances").all()
    rows = [
        {
            "device": d,
            "tco": d.total_cost_of_ownership,
            "replacement_score": d.replacement_recommendation_score,
            "priority": d.replacement_priority_label,
        }
        for d in devices
    ]
    rows.sort(key=lambda r: r["replacement_score"], reverse=True)

    return render(request, "devices/procurement_dashboard.html", {
        "procurement_rows": rows[:50],
        "portfolio_tco": sum((r["tco"] for r in rows), start=0),
        "high_priority_count": sum(1 for r in rows if r["replacement_score"] >= 70),
        "medium_priority_count": sum(1 for r in rows if 40 <= r["replacement_score"] < 70),
    })


# ─── Technician workbench ─────────────────────────────────────────────────────

@login_required
def technician_workbench(request):
    open_work_orders = (
        Maintenance.objects.select_related("device", "device__department")
        .exclude(status__in=["completed", "verified"])
        .order_by("-created_at")[:20]
    )
    return render(request, "devices/technician/workbench.html", {"open_work_orders": open_work_orders})


@login_required
def technician_device_from_qr(request):
    device_id = (request.GET.get("device_id") or "").strip()
    device = get_object_or_404(Device, device_id__iexact=device_id)
    return redirect("technician_device", pk=device.pk)


@login_required
def technician_device(request, pk):
    device = get_object_or_404(Device, pk=pk)
    active_work_order = (
        device.maintenances.exclude(status__in=["completed", "verified"]).order_by("-created_at").first()
    )
    return render(request, "devices/technician/device.html", {
        "device": device,
        "active_work_order": active_work_order,
    })


@login_required
def technician_start_work_order(request, pk):
    device = get_object_or_404(Device, pk=pk)
    if request.method == "POST":
        work_order = Maintenance.objects.create(
            device=device,
            maintenance_type=request.POST.get("maintenance_type", "corrective"),
            technician=request.user.username,
            assigned_technician=request.user.username,
            status="in_progress",
            started_at=timezone.now(),
            description=request.POST.get("description") or "Started from technician mobile flow",
        )
        messages.success(request, f"Work order #{work_order.id} started.")
    return redirect("technician_device", pk=pk)


@login_required
def technician_stop_work_order(request, maintenance_id):
    work_order = get_object_or_404(Maintenance, pk=maintenance_id)
    if request.method == "POST":
        work_order.status = request.POST.get("status", "completed")
        work_order.stopped_at = timezone.now()
        work_order.technician_signature = request.POST.get("technician_signature", request.user.username)
        work_order.notes = request.POST.get("notes", work_order.notes)
        if request.FILES.get("photo_attachment"):
            work_order.photo_attachment = request.FILES["photo_attachment"]
        work_order.save()
        messages.success(request, f"Work order #{work_order.id} updated.")
    return redirect("technician_device", pk=work_order.device_id)


@login_required
def technician_sync_notes(request, maintenance_id):
    work_order = get_object_or_404(Maintenance, pk=maintenance_id)
    if request.method == "POST":
        offline_notes = request.POST.getlist("offline_notes")
        if not offline_notes:
            raw = request.POST.get("offline_notes_blob", "")
            offline_notes = [line.strip() for line in raw.split("\n") if line.strip()]
        for body in offline_notes:
            TechnicianNote.objects.create(
                maintenance=work_order,
                body=body,
                is_offline_created=True,
                synced_at=timezone.now(),
            )
        messages.success(request, f"{len(offline_notes)} offline notes synced.")
    return redirect("technician_device", pk=work_order.device_id)


# ─── HTMX partials ────────────────────────────────────────────────────────────

@require_GET
@login_required
def htmx_stats_strip(request):
    today = date.today()
    active_devices = Device.objects.filter(status="active").count()
    maintenance_devices = Device.objects.filter(status="maintenance").count()
    critical_alerts = Device.objects.filter(next_maintenance__isnull=False).filter(
        Q(next_maintenance__lt=today) | Q(next_maintenance__lte=today + timedelta(days=3))
    ).count()
    return render(request, "partials/stats_strip.html", {
        "active_devices": active_devices,
        "maintenance_devices": maintenance_devices,
        "critical_alerts": critical_alerts,
    })


@require_GET
@login_required
def htmx_todays_maintenance(request):
    today = timezone.now().date()
    todays_maintenance = (
        Maintenance.objects.select_related("device").filter(date=today).order_by("-created_at")[:10]
    )
    return render(request, "partials/todays_maintenance.html", {"todays_maintenance": todays_maintenance})


@require_GET
@login_required
def htmx_critical_zone(request):
    """
    FIX E5: Use a proper ORM queryset with annotation instead of a Python loop.
    This avoids loading all devices into memory.
    """
    from django.db.models import DateField, DurationField, ExpressionWrapper, F, Value
    from django.db.models.functions import Cast

    today = date.today()

    overdue_qs = (
        Device.objects.filter(next_maintenance__lt=today, next_maintenance__isnull=False)
        .only("name", "device_id", "next_maintenance")
        .order_by("next_maintenance")[:5]
    )

    overdue = [
        {
            "name": d.name,
            "device_id": d.device_id,
            "days_overdue": (today - d.next_maintenance).days,
            "next_maintenance": d.next_maintenance,
        }
        for d in overdue_qs
    ]
    return render(request, "partials/critical_zone.html", {"overdue_maintenance": overdue})


# ─── Misc ─────────────────────────────────────────────────────────────────────

@login_required
def team_profile(request):
    return redirect("dashboard")


def set_language_view(request):
    if request.method == "POST":
        language = request.POST.get("language", "en")
        request.session["django_language"] = language
    return redirect(request.META.get("HTTP_REFERER", "/"))


# ─── Health check ─────────────────────────────────────────────────────────────

@require_GET
def healthz(request):
    """Liveness probe — always 200, reports DB connectivity."""
    try:
        connection.ensure_connection()
        db_ok = True
    except Exception:
        db_ok = False
    return JsonResponse({"status": "ok", "db": db_ok}, status=200)
